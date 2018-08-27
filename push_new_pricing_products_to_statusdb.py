#!/usr/bin/env python
"""
Reads the cost_calculator excel sheet and puts all that information
into statusdb.
"""
import argparse
from openpyxl import load_workbook
import coloredlogs
import logging
import yaml
from couchdb import Server
import datetime
from collections import OrderedDict

FIRST_ROW = {'components': 9,
             'products': 4}
SHEET = {'components': 'Price list',
         'products': 'Products'}

# Skip columns which are calculated from the other fields
SKIP = {'components': ['Price', 'Total', 'Per unit'],
        'products': ['Internal', 'External']}

ASSUMED_HEADER_NAMES = ['Category', 'Type', 'Name']

# Assuming the rows are sorted in the preferred order

MAX_NR_ROWS = 200

# Set up a logger with colored output
logger = logging.getLogger(__name__)
logger.propagate = False  # Otherwise the messages appeared twice
coloredlogs.install(level='INFO', logger=logger,
                    fmt='%(asctime)s %(levelname)s %(message)s')


def is_empty_row(comp):
    for k, v in comp.items():
        if v != '':
            return False
    return True


def load_products(wb):
    ws = wb[SHEET['products']]

    row = FIRST_ROW['products']
    header_row = row - 1
    header_cells = ws[header_row]
    header = {}
    for cell in header_cells:
        cell_val = cell.value

        if cell_val not in SKIP['products']:
            # Get cell column as string
            cell_column = cell.coordinate.replace(str(header_row), '')
            header[cell_column] = cell_val

    products = OrderedDict()
    # Unkown number of rows
    while row < MAX_NR_ROWS:
        new_product = {}
        for col, header_val in header.items():
            val = ws["{}{}".format(col, row)].value
            if val is None:
                val = ''
            if header_val == 'Components':
                # Some cells might be interpreted as floats
                # e.g. "37,78"
                val = str(val)
                val = val.replace('.', ',')
                if val:
                    val = [int(prod_id) for prod_id in val.split(',')]
            new_product[header_val] = val

        if not is_empty_row(new_product):
            product_row = row - FIRST_ROW['products'] + 1
            products[product_row] = new_product
        row += 1

    return products


def get_current_version(db):
    view_result = db.view('entire_document/by_version', limit=1,
                          descending=True)
    if view_result.rows:
        return int(view_result.rows[0].value['Version'])
    else:
        return 0


def main(input_sheet, config, user, user_email, push=False):
    with open(config) as settings_file:
        server_settings = yaml.load(settings_file)
    couch = Server(server_settings.get("couch_server", None))

    wb = load_workbook(input_sheet, read_only=True, data_only=True)

    db = couch['pricing_products']
    products = load_products(wb)
    doc = {}
    doc['products'] = products
    doc['Issued by user'] = user
    doc['Issued by user email'] = user_email
    doc['Issued at'] = datetime.datetime.now().isoformat()

    current_version = get_current_version(db)
    doc['Version'] = current_version + 1

    if push:
        db.save(doc)
    else:
        print(doc)


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument('pricing_excel_sheet',
                        help="The excel sheet currently used for pricing")
    parser.add_argument('--statusdb_config', required=True,
                        help='The genomics-status settings.yaml file.')
    parser.add_argument('--push', action='store_true',
                        help='Use this tag to actually push to the databse,'
                        ' otherwise it is just dryrun')
    parser.add_argument('--user', required=True,
                        help='User that change the document')
    parser.add_argument('--user_email', required=True,
                        help='Email used to tell who changed the document')
    args = parser.parse_args()

    main(args.pricing_excel_sheet, args.statusdb_config, args.user,
         args.user_email, push=args.push)
