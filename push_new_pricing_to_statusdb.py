#!/usr/bin/env python
"""
Reads the cost_calculator excel sheet and puts all that information
into statusdb.
"""
import argparse
from openpyxl import load_workbook
import coloredlogs
import logging
import yaml
from couchdb import Server
import datetime
from collections import OrderedDict
import pprint

FIRST_ROW = {'components': 9,
             'products': 4}
SHEET = {'components': 'Price list',
         'products': 'Products'}

# Skip columns which are calculated from the other fields
SKIP = {'components': ['Price', 'Total', 'Per unit'],
        'products': ['Internal', 'External']}

# The name of the _id_ key and which variables that cannot be changed
# while keeping the same _id_. If an update of any of these fields is needed,
# a new id needs to be created.
CONSERVED_KEY_SETS = {'products': ('ID', ['Category', 'Type', 'Name']),
                      'components': ('REF_ID', ['Category', 'Type', 'Product name'])}

# The combination of these "columns" need to be unique within the document
UNIQUE_KEY_SETS = {'products': ('ID', ['Category', 'Type', 'Name']),
                   'components': ('REF_ID', ['Category', 'Type', 'Product name', 'Units'])}

NOT_NULL_KEYS = {'products': ['Category', 'Type', 'Name', 'Re-run fee'],
                 'components': ['Category', 'Type', 'Status',
                                'Product name', 'Units', 'Currency',
                                'List price', 'Discount']}

MAX_NR_ROWS = 200

# Assuming the rows of products are sorted in the preferred order

# Set up a logger with colored output
logger = logging.getLogger('push_new_pricing_logger')
logger.propagate = False  # Otherwise the messages appeared twice
coloredlogs.install(level='INFO', logger=logger,
                    fmt='%(asctime)s %(levelname)s %(message)s')


def check_unique(items, type):
    """Make sure all items within _items_

    fulfill the uniqueness criteria according to the UNIQUE_KEY_SETS
    """
    key_val_set = set()
    for id, item in items.items():
        id_key, keys = UNIQUE_KEY_SETS[type]
        t = tuple(item[key] for key in keys)

        # Check that it is not already added
        if t in key_val_set:
            raise ValueError("Key combination {}:{} is included multiple "
                             "times in the {} sheet. "
                             "ABORTING.".format(keys, t, type))
        key_val_set.add(t)
    return True


def check_conserved(new_items, current_items, type):
    """Ensures the keys in CONSERVED_KEY_SETS are conserved for each given id.

    Compares the new version against the currently active one.
    Params:
        new_items     - A dict of the items that are to be added
                        with ID attribute as the key.
        current_items - A dict of the items currently in the database
                        with ID attribute as the key.
        type          - Either "components" or "products"
    """
    conserved_keys = CONSERVED_KEY_SETS[type][1]

    for id, new_item in new_items.items():
        if str(id) in current_items:
            for conserved_key in conserved_keys:
                if conserved_key not in new_item:
                    raise ValueError("{} column not found in new {} row with "
                                     "id {}. This column needs to be kept "
                                     "conserved. ABORTING!".format(
                                        conserved_key,
                                        type,
                                        id
                                        ))
                if new_item[conserved_key] != current_items[str(id)][conserved_key]:
                    raise ValueError("{} need to be conserved for {}."
                                     " Violated for item with id {}. "
                                     "Found {} for new and {} for current. "
                                     "ABORTING!".format(
                                        conserved_key,
                                        type,
                                        id,
                                        new_item[conserved_key],
                                        current_items[str(id)][conserved_key]
                                        ))
    return True


def check_not_null(items, type):
    """Make sure type specific columns (given by NOT_NULL_KEYS) are not null."""

    not_null_keys = NOT_NULL_KEYS[type]

    for id, item in items.items():
        for not_null_key in not_null_keys:
            if item[not_null_key] is None or item[not_null_key] == '':
                # Special case for discontinued components
                if 'Status' in item and item['Status'] == 'Discontinued':
                    pass
                else:
                    raise ValueError("{} cannot be empty for {}."
                                    " Violated for item with id {}.".\
                                    format(not_null_key, type, id))


def get_current_items(db, type):
    rows = db.view("entire_document/by_version", descending=True, limit=1).rows
    if len(rows) != 0:
        doc = rows[0].value
        return doc[type]
    return {}


def is_empty_row(comp):
    for k, v in comp.items():
        if v != '':
            return False
    return True


def load_products(wb):
    ws = wb[SHEET['products']]

    row = FIRST_ROW['products']
    header_row = row - 1
    header_cells = ws[header_row]
    header = {}
    product_price_columns = {}
    for cell in header_cells:
        cell_val = cell.value
        if cell_val == 'ID':
            cell_val = 'REF_ID'  # Don't want to confuse it with couchdb ids
        # Get cell column as string
        cell_column = cell.coordinate.replace(str(header_row), '')
        if cell_val not in SKIP['products']:
            header[cell_column] = cell_val
        else:
            # save a lookup to find column of prices
            product_price_columns[cell_val] = cell_column

    products = OrderedDict()
    # Unkown number of rows
    while row < MAX_NR_ROWS:
        new_product = {}
        fetch_prices = False # default behaviour
        for col, header_val in header.items():
            val = ws["{}{}".format(col, row)].value
            if val is None:
                val = ''
            if header_val in ['Components', 'Alternative Components']:
                # Some cells might be interpreted as floats
                # e.g. "37,78"
                val = str(val)
                val = val.replace('.', ',')
                if val:
                    val_list = []
                    for comp_id in val.split(','):
                        try:
                            int(comp_id)
                        except ValueError:
                            print("Product on row {} has component with "
                                  "invalid id {}: not an integer, "
                                  " aborting!".format(row, comp_id))
                            raise
                        # Make a list with all individual components
                        val_list.append(comp_id)

                    val = {comp_ref_id: {'quantity': 1} for comp_ref_id in val_list}
                elif header_val == 'Components':
                    # If no components are listed, price should be fetched as well,
                    # unless the row is in fact empty.
                    if not is_empty_row(new_product):
                        fetch_prices = True

            # Comment column occurs after the price columns, so
            # checking for this ensures that the prices have been parsed
            if (header_val == 'Comment') and fetch_prices:
                # Fixed price is added when price does not
                # directly depend on the components
                new_product['fixed_price'] = {}
                int_cell = "{}{}".format(
                                product_price_columns['Internal'],
                                row
                            )
                ext_cell = "{}{}".format(
                                product_price_columns['External'],
                                row
                            )
                new_product['fixed_price']['price_in_sek'] = ws[int_cell].value
                new_product['fixed_price']['external_price_in_sek'] = ws[ext_cell].value

            new_product[header_val] = val

        if not is_empty_row(new_product):
            # The id seems to be stored as a string in the database
            # so might as well always have the ids as strings.
            product_id = str(new_product['REF_ID'])

            # Prepare for a status value on products
            if 'Status' not in new_product:
                new_product['Status'] = "Enabled"

            products[product_id] = new_product
        row += 1

    return products


def load_components(wb):
    ws = wb[SHEET['components']]

    # Unkown number of rows
    row = FIRST_ROW['components']
    header_row = row - 1
    header_cells = ws[header_row]
    header = {}
    for cell in header_cells:
        cell_val = cell.value
        if cell_val == 'ID':
            cell_val = 'REF_ID'  # Don't want to confuse it with couchdb ids
        if cell_val not in SKIP['components']:
            # Get cell column as string
            cell_column = cell.coordinate.replace(str(header_row), '')
            header[cell_column] = cell_val

    components = {}
    while row < MAX_NR_ROWS:
        new_component = {}
        for col, header_val in header.items():
            val = ws["{}{}".format(col, row)].value
            if val is None:
                val = ''
            elif header_val == 'REF_ID':
                # The id seems to be stored as a string in the database
                # so might as well always have the ids as strings.
                try:
                    int(val)
                except ValueError:
                    print("ID value {} for row {} is not an id, "
                          "aborting.".format(val, row))
                val = str(val)

            new_component[header_val] = val

        if new_component['REF_ID'] in components:
            # Violates the uniqueness of the ID
            raise ValueError("ID {} is included multiple "
                             "times in the {} sheet. "
                             "ABORTING.".format(new_component['REF_ID'], type))

        if not is_empty_row(new_component):
            components[new_component['REF_ID']] = new_component
        row += 1

    return components


def get_current_version(db):
    view_result = db.view('entire_document/by_version', limit=1,
                          descending=True)
    if view_result.rows:
        return int(view_result.rows[0].value['Version'])
    else:
        return 0


def compare_two_objects(obj1, obj2, ignore_updated_time=True):
    # Make copies in order to ignore fields
    obj1_copy = obj1.copy()
    obj2_copy = obj2.copy()

    if ignore_updated_time:
        if 'Last Updated' in obj1_copy:
            obj1_copy.pop('Last Updated')
        if 'Last Updated' in obj2_copy:
            obj2_copy.pop('Last Updated')

    return obj1_copy == obj2_copy


def set_last_updated_field(new_objects, current_objects, object_type):
    # if object is not found or changed in current set last updated field
    now = datetime.datetime.now().isoformat()
    for id in new_objects.keys():
        updated = False
        if id in current_objects:
            # Beware! This simple == comparison is quite brittle. Sensitive to
            # str vs int and such.
            the_same = compare_two_objects(new_objects[id],
                                           current_objects[id])
            if not the_same:
                updated = True
        else:
            updated = True

        if updated:
            print("Updating {}: {}".format(object_type, id))
            new_objects[id]['Last Updated'] = now
        else:
            new_objects[id]['Last Updated'] = current_objects[id]['Last Updated']

    return new_objects


def main_push(input_file, config, user, user_email,
              add_components=False, add_products=False, push=False):
    with open(config) as settings_file:
        server_settings = yaml.load(settings_file)
    couch = Server(server_settings.get("couch_server", None))

    wb = load_workbook(input_file, read_only=True, data_only=True)

    # setup a default doc that will be pushed
    doc = {}
    doc['Issued by user'] = user
    doc['Issued by user email'] = user_email
    doc['Issued at'] = datetime.datetime.now().isoformat()

    # A newly pushed document is always a draft
    doc['Draft'] = True

    # --- Components --- #
    comp_db = couch['pricing_components']
    components = load_components(wb)
    check_unique(components, 'components')
    check_not_null(components, 'components')

    current_components = get_current_items(comp_db, 'components')

    if current_components:
        check_conserved(components, current_components, 'components')

    # Modify the `last updated`-field of each item
    components = set_last_updated_field(components,
                                        current_components,
                                        'component')

    # Save it but push it only if products are also parsed correctly
    comp_doc = doc.copy()
    comp_doc['components'] = components

    current_version = get_current_version(comp_db)
    comp_doc['Version'] = current_version + 1

    # --- Products --- #
    prod_db = couch['pricing_products']
    products = load_products(wb)

    check_unique(products, 'products')
    check_not_null(products, 'products')

    current_products = get_current_items(prod_db, 'products')

    if current_products:
        check_conserved(products, current_products, 'products')

    # Modify the `last updated`-field of each item
    products = set_last_updated_field(products,
                                      current_products,
                                      'product')

    prod_doc = doc.copy()
    prod_doc['products'] = products

    current_version = get_current_version(prod_db)
    prod_doc['Version'] = current_version + 1

    # --- Push or Print --- #
    if push:
        comp_db = couch['pricing_components']
        prod_db = couch['pricing_products']

        curr_comp_rows = comp_db.view("entire_document/by_version", descending=True, limit=1).rows
        curr_prod_rows = prod_db.view("entire_document/by_version", descending=True, limit=1).rows

        # Check that the latest one is not a draft
        if (len(curr_comp_rows) == 0) or (len(curr_prod_rows) == 0):
            print("No current version found. This will be the first!")
        else:
            curr_comp_doc = curr_comp_rows[0].value
            curr_prod_doc = curr_prod_rows[0].value
            if curr_comp_doc['Draft'] or curr_prod_doc['Draft']:
                print("Most recent version is a draft. Please remove or "
                      "publish this one before pushing a new draft. Aborting!")
                return

        logger.info(
            'Pushing components document version {}'.format(comp_doc['Version'])
            )
        comp_db.save(comp_doc)

        logger.info(
            'Pushing products document version {}'.format(prod_doc['Version'])
            )
        prod_db.save(prod_doc)
    else:
        # Prettyprint the json output
        pprint.pprint(comp_doc)
        pprint.pprint(prod_doc)

def main_publish(config, user, user_email, dryrun=True):
    with open(config) as settings_file:
        server_settings = yaml.load(settings_file)
    couch = Server(server_settings.get("couch_server", None))

    comp_db = couch['pricing_components']
    prod_db = couch['pricing_products']

    comp_rows = comp_db.view("entire_document/by_version", descending=True, limit=1).rows
    prod_rows = prod_db.view("entire_document/by_version", descending=True, limit=1).rows

    if (len(comp_rows) == 0) or (len(prod_rows) == 0):
        print("No draft version found to publish. Aborting!")
        return

    comp_doc = comp_rows[0].value
    prod_doc = prod_rows[0].value
    if (not comp_doc['Draft']) or (not prod_doc['Draft']):
        print("Most recent version is not a draft. Aborting!")
        return

    now = datetime.datetime.now().isoformat()
    comp_doc['Draft'] = False
    comp_doc['Published'] = now

    prod_doc['Draft'] = False
    prod_doc['Published'] = now

    if not dryrun:
        logger.info(
            'Pushing components document version {}'.format(comp_doc['Version'])
            )
        comp_db.save(comp_doc)

        logger.info(
            'Pushing products document version {}'.format(prod_doc['Version'])
            )
        prod_db.save(prod_doc)
    else:
        print(prod_doc, comp_doc)


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description=__doc__)
    subparsers = parser.add_subparsers(
                            title='actions',
                            dest='subcommand_name',
                            description="Either 'push' for uploading a draft "
                                        "pricing version or 'publish' to make "
                                        "the current draft the latest version"
                                       )

    push_parser = subparsers.add_parser('push')
    push_parser.add_argument('pricing_excel_file',
                        help="The excel file currently used for pricing")
    push_parser.add_argument('--statusdb_config', required=True,
                        help='The genomics-status settings.yaml file.')
    push_parser.add_argument('--push', action='store_true',
                        help='Use this tag to actually push to the databse,'
                        ' otherwise it is just dryrun')
    push_parser.add_argument('--user', required=True,
                        help='User that change the document')
    push_parser.add_argument('--user_email', required=True,
                        help='Email for the user who changed the document')

    publish_parser = subparsers.add_parser('publish')
    publish_parser.add_argument('--statusdb_config', required=True,
                        help='The genomics-status settings.yaml file.')
    publish_parser.add_argument('--user', required=True,
                        help='User that change the document')
    publish_parser.add_argument('--user_email', required=True,
                        help='Email for the user who changed the document')
    publish_parser.add_argument('--dryrun', action='store_true',
                                help="Use this tag to only print what would "
                                "have been done")

    args = parser.parse_args()

    if args.subcommand_name == 'push':
        main_push(args.pricing_excel_file, args.statusdb_config, args.user,
                  args.user_email, push=args.push)
    elif args.subcommand_name == 'publish':
        main_publish(args.statusdb_config, args.user, args.user_email,
                     dryrun=args.dryrun)
